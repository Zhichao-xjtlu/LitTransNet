#!/usr/bin/env python3
"""Prepare a local literature corpus for retrieval without OCR or network access."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Iterator


SUPPORTED_SUFFIXES = {".pdf", ".txt", ".csv", ".xlsx"}
PROJECT_ROOT = Path(__file__).resolve().parents[2]


@dataclass
class FileResult:
    source_file: str
    file_type: str
    chunk_count: int = 0
    parse_status: str = "success"
    error_message: str = ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert local PDF, TXT, CSV, and XLSX literature into JSONL chunks."
    )
    parser.add_argument("--literature_dir", default="literature", help="Input directory (default: literature)")
    parser.add_argument(
        "--output_jsonl",
        default="rag/corpus/chunks.jsonl",
        help="Output corpus path (default: rag/corpus/chunks.jsonl)",
    )
    parser.add_argument("--chunk_size", type=int, default=1200, help="Maximum chunk size in characters")
    parser.add_argument("--chunk_overlap", type=int, default=200, help="Overlap between long text chunks")
    parser.add_argument(
        "--reports_dir",
        default="",
        help="Report directory. Default: sibling reports/ beside the output corpus/ directory.",
    )
    args = parser.parse_args()
    if args.chunk_size <= 0:
        parser.error("--chunk_size must be greater than zero")
    if args.chunk_overlap < 0:
        parser.error("--chunk_overlap cannot be negative")
    if args.chunk_overlap >= args.chunk_size:
        parser.error("--chunk_overlap must be smaller than --chunk_size")
    return args


def resolve_from_project(path_text: str) -> Path:
    path = Path(path_text).expanduser()
    return path if path.is_absolute() else PROJECT_ROOT / path


def stable_chunk_id(
    source_file: str,
    page: int | None,
    sheet_name: str | None,
    row_index: int | None,
    chunk_number: int,
) -> str:
    locator = f"{source_file}|p={page}|s={sheet_name}|r={row_index}|c={chunk_number}"
    digest = hashlib.sha256(locator.encode("utf-8")).hexdigest()[:20]
    return f"chunk_{digest}"


def split_text(text: str, chunk_size: int, overlap: int) -> Iterator[tuple[int, int, str]]:
    """Yield character-based chunks, preferring whitespace boundaries."""
    if not text:
        return
    length = len(text)
    start = 0
    while start < length:
        end = min(start + chunk_size, length)
        if end < length:
            minimum_break = start + max(1, int(chunk_size * 0.7))
            whitespace = max(text.rfind(" ", minimum_break, end), text.rfind("\n", minimum_break, end))
            if whitespace > start:
                end = whitespace

        content_start = start
        content_end = end
        while content_start < content_end and text[content_start].isspace():
            content_start += 1
        while content_end > content_start and text[content_end - 1].isspace():
            content_end -= 1
        if content_start < content_end:
            yield content_start, content_end, text[content_start:content_end]
        if end >= length:
            break
        next_start = max(end - overlap, start + 1)
        start = next_start


def looks_like_heading(line: str) -> bool:
    stripped = line.strip()
    if not stripped or len(stripped) > 120 or stripped.endswith((".", ";", ",")):
        return False
    if re.match(r"^(?:\d+(?:\.\d+)*|[A-Z])[\s.)-]+\S+", stripped):
        return True
    words = stripped.split()
    if not (1 <= len(words) <= 12) or not any(char.isalpha() for char in stripped):
        return False
    return stripped.isupper() or sum(word[:1].isupper() for word in words) >= max(1, len(words) - 1)


def section_at(text: str, char_start: int) -> str | None:
    section: str | None = None
    position = 0
    for line in text.splitlines(keepends=True):
        if position > char_start:
            break
        if looks_like_heading(line):
            section = line.strip()
        position += len(line)
    return section


def extract_pdf_pages(pdf_path: Path) -> tuple[list[str], str]:
    """Extract PDF text pages from the embedded text layer only."""
    reader_class = None
    backend = ""
    try:
        from pypdf import PdfReader

        reader_class = PdfReader
        backend = "pypdf"
    except ImportError:
        try:
            from PyPDF2 import PdfReader

            reader_class = PdfReader
            backend = "PyPDF2"
        except ImportError:
            pass

    if reader_class is not None:
        reader = reader_class(str(pdf_path))
        return [(page.extract_text() or "") for page in reader.pages], backend

    executable = shutil.which("pdftotext")
    if executable:
        completed = subprocess.run(
            [executable, "-enc", "UTF-8", "-layout", str(pdf_path), "-"],
            check=False,
            capture_output=True,
            timeout=180,
        )
        if completed.returncode != 0:
            error = completed.stderr.decode("utf-8", errors="replace").strip()
            raise RuntimeError(f"pdftotext exited with code {completed.returncode}: {error}")
        text = completed.stdout.decode("utf-8", errors="replace")
        pages = text.split("\f")
        if pages and not pages[-1].strip():
            pages.pop()
        return pages, "pdftotext"

    raise RuntimeError(
        "No PDF text parser is available. Install pypdf (`pip install pypdf`) "
        "or make pdftotext available. OCR is intentionally not used."
    )


def make_chunk(
    *,
    source_file: str,
    file_type: str,
    text: str,
    char_start: int,
    char_end: int,
    chunk_number: int,
    page: int | None = None,
    sheet_name: str | None = None,
    row_index: int | None = None,
    section: str | None = None,
) -> dict[str, Any]:
    return {
        "chunk_id": stable_chunk_id(source_file, page, sheet_name, row_index, chunk_number),
        "source_file": source_file,
        "file_type": file_type,
        "page": page,
        "sheet_name": sheet_name,
        "row_index": row_index,
        "section": section,
        "text": text,
        "char_start": char_start,
        "char_end": char_end,
    }


def chunks_from_pdf(
    path: Path, source_file: str, chunk_size: int, overlap: int
) -> tuple[list[dict[str, Any]], str]:
    pages, backend = extract_pdf_pages(path)
    chunks: list[dict[str, Any]] = []
    for page_number, page_text in enumerate(pages, start=1):
        for number, (start, end, text) in enumerate(split_text(page_text, chunk_size, overlap)):
            chunks.append(
                make_chunk(
                    source_file=source_file,
                    file_type="pdf",
                    page=page_number,
                    sheet_name=None,
                    row_index=None,
                    section=section_at(page_text, start),
                    text=text,
                    char_start=start,
                    char_end=end,
                    chunk_number=number,
                )
            )
    return chunks, backend


def chunks_from_txt(path: Path, source_file: str, chunk_size: int, overlap: int) -> list[dict[str, Any]]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    return [
        make_chunk(
            source_file=source_file,
            file_type="txt",
            page=None,
            sheet_name=None,
            row_index=None,
            section=section_at(text, start),
            text=content,
            char_start=start,
            char_end=end,
            chunk_number=number,
        )
        for number, (start, end, content) in enumerate(split_text(text, chunk_size, overlap))
    ]


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def serialize_row(headers: list[str], values: Iterable[Any]) -> str:
    pairs: list[str] = []
    for index, value in enumerate(values):
        rendered = cell_text(value)
        if not rendered:
            continue
        header = headers[index] if index < len(headers) and headers[index] else f"column_{index + 1}"
        pairs.append(f"{header}: {rendered}")
    return " | ".join(pairs)


def chunks_from_rows(
    rows: Iterable[tuple[int, list[Any]]],
    headers: list[str],
    source_file: str,
    file_type: str,
    sheet_name: str | None,
    chunk_size: int,
    overlap: int,
) -> list[dict[str, Any]]:
    chunks: list[dict[str, Any]] = []
    for row_index, values in rows:
        row_text = serialize_row(headers, values)
        if not row_text:
            continue
        for number, (start, end, content) in enumerate(split_text(row_text, chunk_size, overlap)):
            chunks.append(
                make_chunk(
                    source_file=source_file,
                    file_type=file_type,
                    page=None,
                    sheet_name=sheet_name,
                    row_index=row_index,
                    section=None,
                    text=content,
                    char_start=start,
                    char_end=end,
                    chunk_number=number,
                )
            )
    return chunks


def chunks_from_csv(path: Path, source_file: str, chunk_size: int, overlap: int) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
        reader = csv.reader(handle)
        try:
            headers = [cell_text(value) for value in next(reader)]
        except StopIteration:
            return []
        rows = ((row_number, list(values)) for row_number, values in enumerate(reader, start=2))
        return chunks_from_rows(rows, headers, source_file, "csv", None, chunk_size, overlap)


def chunks_from_xlsx(path: Path, source_file: str, chunk_size: int, overlap: int) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("XLSX support requires openpyxl (`pip install openpyxl`).") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    chunks: list[dict[str, Any]] = []
    try:
        for worksheet in workbook.worksheets:
            iterator = worksheet.iter_rows(values_only=True)
            try:
                headers = [cell_text(value) for value in next(iterator)]
            except StopIteration:
                continue
            rows = ((row_number, list(values)) for row_number, values in enumerate(iterator, start=2))
            chunks.extend(
                chunks_from_rows(
                    rows, headers, source_file, "xlsx", worksheet.title, chunk_size, overlap
                )
            )
    finally:
        workbook.close()
    return chunks


def write_csv(path: Path, fieldnames: list[str], rows: Iterable[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def prepare_corpus(
    literature_dir: Path, output_jsonl: Path, chunk_size: int, overlap: int
) -> tuple[int, list[FileResult]]:
    if not literature_dir.is_dir():
        raise FileNotFoundError(f"Literature directory does not exist: {literature_dir}")

    files = sorted(
        (path for path in literature_dir.rglob("*") if path.is_file() and path.suffix.lower() in SUPPORTED_SUFFIXES),
        key=lambda item: item.relative_to(literature_dir).as_posix().casefold(),
    )
    output_jsonl.parent.mkdir(parents=True, exist_ok=True)
    results: list[FileResult] = []
    total_chunks = 0

    with output_jsonl.open("w", encoding="utf-8", newline="\n") as output:
        for path in files:
            source_file = path.relative_to(literature_dir).as_posix()
            suffix = path.suffix.lower()
            result = FileResult(source_file=source_file, file_type=suffix.lstrip("."))
            try:
                backend = ""
                if suffix == ".pdf":
                    chunks, backend = chunks_from_pdf(path, source_file, chunk_size, overlap)
                elif suffix == ".txt":
                    chunks = chunks_from_txt(path, source_file, chunk_size, overlap)
                elif suffix == ".csv":
                    chunks = chunks_from_csv(path, source_file, chunk_size, overlap)
                else:
                    chunks = chunks_from_xlsx(path, source_file, chunk_size, overlap)

                result.chunk_count = len(chunks)
                if not chunks:
                    result.parse_status = "empty"
                    result.error_message = "No extractable text found (OCR is disabled)."
                elif backend:
                    result.parse_status = f"success:{backend}"
                for chunk in chunks:
                    output.write(json.dumps(chunk, ensure_ascii=False) + "\n")
                total_chunks += len(chunks)
            except Exception as exc:  # Continue so that one bad document cannot abort the corpus.
                result.parse_status = "failed"
                result.error_message = f"{type(exc).__name__}: {exc}"
            results.append(result)

    return total_chunks, results


def main() -> int:
    args = parse_args()
    literature_dir = resolve_from_project(args.literature_dir)
    output_jsonl = resolve_from_project(args.output_jsonl)
    reports_dir = (
        resolve_from_project(args.reports_dir)
        if args.reports_dir
        else output_jsonl.parent.parent / "reports"
    )
    reports_dir.mkdir(parents=True, exist_ok=True)
    try:
        total_chunks, results = prepare_corpus(
            literature_dir, output_jsonl, args.chunk_size, args.chunk_overlap
        )
    except (OSError, ValueError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    summary_rows = [
        {
            "source_file": item.source_file,
            "file_type": item.file_type,
            "chunk_count": item.chunk_count,
            "parse_status": item.parse_status,
            "error_message": item.error_message,
        }
        for item in results
    ]
    write_csv(
        reports_dir / "corpus_summary.csv",
        ["source_file", "file_type", "chunk_count", "parse_status", "error_message"],
        summary_rows,
    )
    failure_rows = [
        row for row in summary_rows if row["parse_status"] in {"failed", "empty"}
    ]
    write_csv(
        reports_dir / "parse_failures.csv",
        ["source_file", "file_type", "chunk_count", "parse_status", "error_message"],
        failure_rows,
    )
    print(f"Processed {len(results)} files and wrote {total_chunks} chunks to {output_jsonl}")
    print(f"Parse failures: {len(failure_rows)}; reports: {reports_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
